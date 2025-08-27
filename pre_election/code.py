import os
import pandas as pd
import requests
from io import BytesIO
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import random

# === CONFIG ===
INPUT_FILE = "pre_election_voters (2).xlsx"
OUTPUT_FILE = "pre_election_voters_marked.xlsx"
REPORT_FILE = "voter_report.pdf"
IMAGE_FOLDER = "voters"

os.makedirs(IMAGE_FOLDER, exist_ok=True)

# === Load Excel ===
df = pd.read_excel(INPUT_FILE)

# Ensure Duplicate column exists
if "Duplicate" not in df.columns:
    df["Duplicate"] = False

# === Detect duplicates ===
df["Duplicate"] = df.duplicated(subset=["Voter Name"], keep=False)

# === Function to download face from randomuser.me ===
def download_image(voter_name, save_path):
    gender = random.choice(["men", "women"])  # randomly assign gender
    face_id = random.randint(1, 99)           # random face id
    url = f"https://randomuser.me/api/portraits/{gender}/{face_id}.jpg"
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        img = Image.open(BytesIO(response.content)).convert("RGB")
        img.save(save_path, "JPEG")
    except Exception as e:
        print(f"⚠️ Failed to download image for {voter_name}: {e}")

# === Attach images to voters ===
image_files = []
for idx, row in df.iterrows():
    voter_name = row["Voter Name"]
    img_filename = f"{voter_name.replace(' ', '_')}.jpg"
    img_path = os.path.join(IMAGE_FOLDER, img_filename)

    if not os.path.exists(img_path):
        download_image(voter_name, img_path)

    image_files.append(img_filename)

df["Image"] = image_files

# === Save Excel with duplicates column ===
df.to_excel(OUTPUT_FILE, index=False)

# === Color duplicates in Excel ===
wb = load_workbook(OUTPUT_FILE)
ws = wb.active
dup_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # red
unique_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")  # green

dup_col = list(df.columns).index("Duplicate") + 1
for row in range(2, ws.max_row + 1):  # skip header
    cell = ws.cell(row=row, column=dup_col)
    if cell.value is True:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = dup_fill
    else:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = unique_fill

wb.save(OUTPUT_FILE)

# === Generate PDF Report ===
styles = getSampleStyleSheet()
doc = SimpleDocTemplate(REPORT_FILE, pagesize=A4)
story = []

story.append(Paragraph("<b>Voter List Report</b>", styles["Title"]))
story.append(Spacer(1, 12))

total_voters = len(df)
duplicates = df["Duplicate"].sum()
unique = total_voters - duplicates

story.append(Paragraph(f"Total Voters: {total_voters}", styles["Normal"]))
story.append(Paragraph(f"Unique Voters: {unique}", styles["Normal"]))
story.append(Paragraph(f"Duplicate Voters: {duplicates}", styles["Normal"]))
story.append(Spacer(1, 24))

# Removed the sample images section

doc.build(story)

print(f"✅ Processing complete!\nExcel saved as: {OUTPUT_FILE}\nPDF report saved as: {REPORT_FILE}")
